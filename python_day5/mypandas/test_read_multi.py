import openpyxl
book = openpyxl.load_workbook('weather.xlsx')
sheet = book.active
cells = sheet['A1': 'D19'] #drag
for c1, c2, c3, c4 in cells:
    print("%s\t%s\t%s\t%s" % ( c1.value, c2.value, c3.value, c4.value))

print ( sheet.max_column, sheet.max_row )

cells = sheet['B3':'B5']
for c1, in cells:
# for c1 in cells:
    print("%s\t" % (c1.value))