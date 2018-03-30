from openpyxl import Workbook
from bs4 import BeautifulSoup
import urllib.request as req

book = Workbook()
sheet = book.active
sheet['A1'] = 1
sheet.cell(row=2, column=2).value = 2
sheet.cell(row=3, column=2).value = 2
book.save('write2cell.xlsx')

print('end')


url = 'http://www.weather.go.kr/wid/queryDFSRSS.jsp?zone=1150060300'
response = req.urlopen(url)
soup = BeautifulSoup(response, "html.parser")

datas = soup.select('data')
book = Workbook()
sheet = book.active
# print(datas)
for data in datas:
#     print(data.hour, data.hour.name, data.hour.string, data.hour.parent.name, data['seq'])
    hour = int(data.hour.string)
    day = int(data.day.string)
    temp = data.temp.string
    wfkor = data.wfkor.string
    sheet.append((hour,day,temp,wfkor))

    
book.save('weather.xlsx')

print('end')